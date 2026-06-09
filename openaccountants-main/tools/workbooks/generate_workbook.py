#!/usr/bin/env python3
"""
OpenAccountants — Skill Verification Workbook generator.

Reads a jurisdiction JSON (data/<code>.json) describing skills -> line items and
renders <CODE>_Tax_Skills_Verification.xlsx in the "flag-only-what's-wrong"
format:

  * The accountant marks ONLY incorrect rows (with a reason). Anything left
    blank is taken as CONFIRMED CORRECT — there is deliberately no "Correct"
    option (that's the change from the old Canada template).
  * A "Missing" block at the bottom of each tab invites additions, credited to
    the verifier on the published skill.

Usage:
    uv run --with openpyxl python generate_workbook.py data/us-fl.json
    python generate_workbook.py data/us-fl.json out/

A hidden `_map` sheet records tab -> {slug, version} so a future parser can map
a returned workbook back to skills (the round-trip in WORKBOOK_GENERATOR_PLAN).
"""
import json
import re
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

# Flag dropdown — NO "Correct": blank = correct by default.
FLAG_OPTIONS = ["Wrong", "Outdated", "Missing detail", "Unsure"]
HEADERS = [
    "Section / Location",
    "What the skill says now",
    "Flag (blank = correct)",
    "Correct value / what it should say",
    "Why / reason",
    "Reference / source",
]
COL_W = [34, 52, 20, 34, 34, 26]

NAVY, SECTION, FLAGCOL, MISSING = "1F3A5F", "D6E4F0", "FFF7E6", "FDECEC"
HEAD_FONT = Font(bold=True, color="FFFFFF", size=11)
SECTION_FONT = Font(bold=True, color="1F3A5F", size=11)
_thin = Side(style="thin", color="C9D4E0")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
WRAP_TOP = Alignment(vertical="top", wrap_text=True)


def _fill(color):
    return PatternFill("solid", fgColor=color)


# Excel sheet titles cannot contain : \ / ? * [ ] and max 31 chars.
_INVALID_TITLE = re.compile(r"[:\\/?*\[\]]")


def _sheet_title(name):
    return _INVALID_TITLE.sub("-", name)[:31]


def _style_header(ws):
    for c, label in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=c, value=label)
        cell.font = HEAD_FONT
        cell.fill = _fill(NAVY)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[1].height = 30
    for i, w in enumerate(COL_W, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


def _add_skill_tab(wb, skill):
    ws = wb.create_sheet(_sheet_title(skill["tab"]))
    _style_header(ws)
    dv = DataValidation(
        type="list", formula1='"%s"' % ",".join(FLAG_OPTIONS), allow_blank=True
    )
    ws.add_data_validation(dv)

    r = 2
    for section in skill["sections"]:
        ws.cell(row=r, column=1, value=section["section"]).font = SECTION_FONT
        for c in range(1, len(HEADERS) + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = _fill(SECTION)
            cell.border = BORDER
        r += 1
        for row in section["rows"]:
            ws.cell(row=r, column=1, value=row.get("item", ""))
            ws.cell(row=r, column=2, value=row.get("value", ""))
            ws.cell(row=r, column=6, value=row.get("ref", ""))
            for c in range(1, len(HEADERS) + 1):
                cell = ws.cell(row=r, column=c)
                cell.alignment = WRAP_TOP
                cell.border = BORDER
            ws.cell(row=r, column=3).fill = _fill(FLAGCOL)
            dv.add(ws.cell(row=r, column=3))
            r += 1

    # "Missing" block — add what we don't cover (credited to the verifier).
    r += 1
    head = ws.cell(
        row=r,
        column=1,
        value="➕ Missing — add anything we don't cover (you'll be credited as verifier)",
    )
    head.font = Font(bold=True, color="B00020", size=11)
    head.fill = _fill(MISSING)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    r += 1
    note = ws.cell(
        row=r,
        column=1,
        value="Add a rate, threshold, rule, deadline, or credit we're missing — one per row "
        "(put it in 'Correct value', why in 'Why / reason', source in 'Reference'). "
        "Every addition publishes your name + credential on this skill at openaccountants.com.",
    )
    note.font = Font(italic=True, color="666666", size=9)
    note.alignment = WRAP_TOP
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    r += 1
    for _ in range(8):
        for c in range(1, len(HEADERS) + 1):
            ws.cell(row=r, column=c).border = BORDER
        ws.cell(row=r, column=3).fill = _fill(FLAGCOL)
        dv.add(ws.cell(row=r, column=3))
        r += 1
    return ws.title


def _instructions(wb, data):
    ws = wb.create_sheet("Instructions", 0)
    ws.column_dimensions["A"].width = 120
    block = [
        ("%s Tax Skills — Verification Workbook" % data["jurisdiction_label"],
         Font(bold=True, size=16, color="1F3A5F")),
        ("OpenAccountants — openaccountants.com", Font(size=10, color="666666")),
        ("", None),
        ("How this works — please read", Font(bold=True, size=12)),
        ("• You only mark what's WRONG or MISSING. Anything you leave blank, we take as CONFIRMED CORRECT.", None),
        ("• For a wrong row: set the Flag (Wrong / Outdated / Missing detail / Unsure), put the right value in "
         "'Correct value', and a short 'Why / reason'.", None),
        ("• 'Reference / source' is pre-filled with the skill's citation — correct it if it's wrong.", None),
        ("• Found something we don't cover? Add it in the '➕ Missing' block at the bottom of each tab.", None),
        ("", None),
        ("Why bother — you get credited", Font(bold=True, size=12)),
        ("Every correction or addition you make publishes your name and credential as the verifying accountant on "
         "the skill at openaccountants.com — the skill your clients' AI then follows, with your booking link attached.", None),
        ("", None),
        ("Skills in this workbook (one per tab):", Font(bold=True, size=11)),
    ]
    r = 1
    for text, font in block:
        cell = ws.cell(row=r, column=1, value=text)
        if font:
            cell.font = font
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        r += 1
    for s in data["skills"]:
        ws.cell(row=r, column=1, value="   • %s  (%s)" % (s["tab"], s.get("slug", "")))
        r += 1
    r += 1
    ws.cell(row=r, column=1,
            value="Jurisdiction: %s   |   Tax year: %s" % (data["jurisdiction_code"], data.get("tax_year", "")))
    r += 1
    ws.cell(row=r, column=1,
            value="Verifier (print name + credential): ____________________________     Date: ____________")


def _hidden_map(wb, data):
    ws = wb.create_sheet("_map")
    ws.sheet_state = "hidden"
    ws.append(["tab", "slug", "version"])
    for s in data["skills"]:
        ws.append([_sheet_title(s["tab"]), s.get("slug", ""), str(s.get("version", ""))])


def build(data):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for skill in data["skills"]:
        _add_skill_tab(wb, skill)
    _instructions(wb, data)   # inserted at index 0 -> first tab
    _hidden_map(wb, data)
    return wb


def main():
    if len(sys.argv) < 2:
        print("usage: generate_workbook.py <data.json> [out_dir]")
        sys.exit(1)
    data_path = Path(sys.argv[1])
    out_dir = Path(sys.argv[2]) if len(sys.argv) > 2 else data_path.parent.parent / "out"
    out_dir.mkdir(parents=True, exist_ok=True)
    data = json.loads(data_path.read_text())
    wb = build(data)
    out = out_dir / ("%s_Tax_Skills_Verification.xlsx" % data["jurisdiction_code"])
    wb.save(out)
    n_rows = sum(len(sec["rows"]) for s in data["skills"] for sec in s["sections"])
    print("Wrote %s — %d skill tabs, %d pre-filled line items" % (out, len(data["skills"]), n_rows))


if __name__ == "__main__":
    main()
