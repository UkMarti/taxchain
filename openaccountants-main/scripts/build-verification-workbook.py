#!/usr/bin/env python3
"""
Build a verifier-friendly xlsx workbook for a credentialed accountant to
review OpenAccountants skills offline. They open the file, work through each
skill sheet, mark every section as correct / wrong / needs-change, write
corrections inline, and sign off at the bottom of each skill. They email
the file back. We parse it and apply their corrections to Supabase + register
them as the named verifier.

Usage:
    # By explicit slug list (most common — review a curated set)
    python3 scripts/build-verification-workbook.py \\
        --slugs us-gaap-asc606-revenue,ifrs15-revenue,... \\
        --verifier "Dr. Andrei Belonogov" \\
        --output /tmp/oa-verification-andrei.xlsx

    # By jurisdiction (review everything for one state / country)
    python3 scripts/build-verification-workbook.py \\
        --jurisdiction US-ND \\
        --verifier "ND CPA Name" \\
        --output /tmp/oa-verification-nd.xlsx

Env required:
    NEXT_PUBLIC_SUPABASE_URL
    SUPABASE_SERVICE_ROLE_KEY
"""

import argparse
import json
import os
import re
import sys
import urllib.error
import urllib.request
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


# ---------------------------------------------------------------------------
# Args + env
# ---------------------------------------------------------------------------

ap = argparse.ArgumentParser()
ap.add_argument("--slugs", help="Comma-separated skill slugs to include.")
ap.add_argument("--jurisdiction", help="Jurisdiction code (e.g. US-ND, MT).")
ap.add_argument("--verifier", required=True, help="Verifier display name (goes on cover sheet + per-skill sign-off rows).")
ap.add_argument("--credential", default="", help="Verifier credential (e.g. 'CPA', 'CA(SA)').")
ap.add_argument("--title", default="OpenAccountants — Skill Verification Workbook", help="Workbook title shown on the cover.")
ap.add_argument("--output", required=True, help="Output .xlsx path.")
args = ap.parse_args()

if not args.slugs and not args.jurisdiction:
    sys.exit("Provide --slugs OR --jurisdiction.")

URL = os.environ["NEXT_PUBLIC_SUPABASE_URL"]
KEY = os.environ["SUPABASE_SERVICE_ROLE_KEY"]
HEADERS = {"apikey": KEY, "Authorization": f"Bearer {KEY}"}


# ---------------------------------------------------------------------------
# Supabase REST helpers
# ---------------------------------------------------------------------------

def http_get(path: str):
    req = urllib.request.Request(f"{URL}{path}", headers=HEADERS)
    with urllib.request.urlopen(req) as r:
        return json.load(r)


def fetch_skills():
    """Return list of skill rows with markdown_content joined from current version."""
    if args.slugs:
        slug_list = [s.strip() for s in args.slugs.split(",") if s.strip()]
        in_clause = "(" + ",".join(f'"{s}"' for s in slug_list) + ")"
        skills = http_get(
            f"/rest/v1/skills?slug=in.{in_clause}&is_published=eq.true"
            "&select=id,slug,name,description,tier,category,jurisdiction,tax_year&order=slug"
        )
    else:
        skills = http_get(
            f"/rest/v1/skills?jurisdiction=eq.{args.jurisdiction}&is_published=eq.true"
            "&select=id,slug,name,description,tier,category,jurisdiction,tax_year&order=slug"
        )

    for s in skills:
        versions = http_get(
            f"/rest/v1/skill_versions?skill_id=eq.{s['id']}&is_current=eq.true"
            "&select=markdown_content,version&limit=1"
        )
        s["markdown"] = versions[0]["markdown_content"] if versions else ""
        s["version"] = versions[0]["version"] if versions else None
    return skills


# ---------------------------------------------------------------------------
# Section parsing
# ---------------------------------------------------------------------------

def strip_frontmatter(md: str) -> str:
    m = re.match(r"^---\s*\n.+?\n---\s*\n", md, re.S)
    return md[m.end():] if m else md


def parse_sections(md: str):
    """Return list of (heading_level, heading_text, body_markdown).

    Splits on level-2+ ATX headings (## / ### / ...). Content before the first
    heading becomes a synthetic "Preamble" section.
    """
    body = strip_frontmatter(md)
    parts = re.split(r"(?m)^(#{2,6}\s+.+?)$", body)
    # parts is [preamble, h1, body1, h2, body2, ...] — odd indices are headings.
    sections = []
    preamble = parts[0].strip()
    if preamble:
        sections.append((1, "Preamble (above first ## heading)", preamble))
    for i in range(1, len(parts), 2):
        heading_raw = parts[i].strip()
        level = len(re.match(r"^#+", heading_raw).group(0))
        heading = re.sub(r"^#+\s*", "", heading_raw).strip()
        content = parts[i + 1].strip() if i + 1 < len(parts) else ""
        sections.append((level, heading, content))
    return sections


# ---------------------------------------------------------------------------
# xlsx styling helpers
# ---------------------------------------------------------------------------

THICK = Side(border_style="medium", color="000000")
THIN = Side(border_style="thin", color="999999")
HEADER_FILL = PatternFill("solid", fgColor="047857")
SUBHEAD_FILL = PatternFill("solid", fgColor="ECFDF5")
SIGNOFF_FILL = PatternFill("solid", fgColor="FFF7ED")
ZEBRA_FILL = PatternFill("solid", fgColor="F9FAFB")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=18)
H2_FONT = Font(bold=True, size=13)
WRAP = Alignment(wrap_text=True, vertical="top")
TOP = Alignment(vertical="top")

STATUS_OPTIONS = '"correct,needs-correction,wrong,needs-context,skip"'
STATUS_DV = DataValidation(type="list", formula1=STATUS_OPTIONS, allow_blank=True)
STATUS_DV.error = "Pick one of: correct / needs-correction / wrong / needs-context / skip"
STATUS_DV.prompt = "correct = fact is right • needs-correction = small fix (write it in the next column) • wrong = section is misleading • needs-context = unclear / add detail • skip = not in your scope"
STATUS_DV.promptTitle = "Section status"


def sheet_name(slug: str) -> str:
    """Excel sheet name max 31 chars, no [], /, \\, ?, *, :."""
    name = re.sub(r'[\[\]/\\?*:]', '-', slug)
    if len(name) > 31:
        # Drop "us-gaap-" / "ifrs" prefix to fit
        for prefix in ("us-gaap-", "ifrs-", "financial-reporting-"):
            if name.startswith(prefix):
                trimmed = name[len(prefix):]
                if len(trimmed) <= 31:
                    return trimmed
                name = trimmed
                break
        if len(name) > 31:
            name = name[:31]
    return name


# ---------------------------------------------------------------------------
# Build workbook
# ---------------------------------------------------------------------------

def main():
    skills = fetch_skills()
    if not skills:
        sys.exit("No skills returned. Check slug list or jurisdiction.")

    print(f"Fetched {len(skills)} skill(s) from Supabase.")
    for s in skills:
        n_sections = len(parse_sections(s["markdown"]))
        print(f"  - {s['slug']:50s} sections={n_sections} jur={s['jurisdiction']}")

    wb = Workbook()
    wb.remove(wb.active)

    build_cover_sheet(wb, skills)
    build_instructions_sheet(wb)
    build_summary_sheet(wb, skills)
    for s in skills:
        build_skill_sheet(wb, s)

    wb.save(args.output)
    print(f"\nWrote {args.output} ({os.path.getsize(args.output) / 1024:.1f} KB)")


def build_cover_sheet(wb, skills):
    ws = wb.create_sheet("Cover", 0)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 90

    ws["A1"] = "OpenAccountants"
    ws["A1"].font = Font(bold=True, color="047857", size=24)
    ws.row_dimensions[1].height = 32

    ws["A2"] = args.title
    ws["A2"].font = Font(bold=True, size=14, color="111111")
    ws.merge_cells("A2:B2")

    fields = [
        ("Verifier", args.verifier),
        ("Credential", args.credential or "(fill in below)"),
        ("Date prepared", date.today().isoformat()),
        ("Skills in scope", str(len(skills))),
        ("MCP endpoint", "https://www.openaccountants.com/api/mcp"),
        ("Network page", "https://www.openaccountants.com/network"),
        ("Contact", "info@openaccountants.com"),
    ]
    for i, (k, v) in enumerate(fields, start=4):
        ws.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws.cell(row=i, column=2, value=v)

    row = 4 + len(fields) + 1
    ws.cell(row=row, column=1, value="How to use this workbook").font = H2_FONT
    ws.cell(row=row, column=1).fill = SUBHEAD_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)

    notes = [
        "1. Skip to the 'Instructions' tab for the full how-to.",
        "2. The 'Summary' tab lists every skill in scope with a 'Your sign-off' column. Mark each row as you finish that skill.",
        "3. Each remaining tab is one skill, broken into review-friendly sections. Read each section and pick a status from the dropdown.",
        "4. When you have corrections, write them inline. Don't worry about formatting — plain English is fine.",
        "5. At the bottom of every skill sheet there's an 'Overall sign-off' block. Fill it in before moving on.",
        "6. Save the file. Email it back to info@openaccountants.com.",
        "7. We apply your corrections to the live MCP, list you on the verifier wall at openaccountants.com/network, and route any future client review request for this scope to your Calendly.",
        "",
        "Honesty rule: if you didn't read a section, mark it 'skip'. We never claim someone verified a skill they didn't actually look at.",
    ]
    for n in notes:
        row += 1
        ws.cell(row=row, column=1, value=n).alignment = WRAP
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)


def build_instructions_sheet(wb):
    ws = wb.create_sheet("Instructions", 1)
    ws.column_dimensions["A"].width = 100

    ws["A1"] = "Instructions"
    ws["A1"].font = Font(bold=True, size=18)
    rows = [
        "",
        "What is OpenAccountants?",
        "An open-source library of accounting and tax computation skills that AI agents (Claude, ChatGPT, Cursor, etc.) follow instead of hallucinating. Each skill carries rates, thresholds, references to the actual statutes / standards, worked examples, and audit flash points. Skills are served via our MCP server at openaccountants.com/api/mcp.",
        "",
        "What does 'verified' mean here?",
        "Tier 1 skills carry a named licensed practitioner's name and credential. When an AI invokes our request_accountant_review tool, we route the user to the named verifier for that scope. Your name and Calendly appear on every output. There is no anonymous 'verified by our team' shield — only real, named, credentialed people.",
        "",
        "How to mark each section",
        "Each skill sheet has a 'Section' column on the left, 'Section content' next to it, then a 'Status' dropdown with these options:",
        "  • correct — the section is accurate as written.",
        "  • needs-correction — a small, specific fact is wrong. Write the correction in the 'Correction / note' column.",
        "  • wrong — the section is materially misleading or hallucinated. Mark it red; write what should replace it.",
        "  • needs-context — the section is accurate but incomplete; suggest what to add.",
        "  • skip — out of your scope or you didn't have time to assess.",
        "",
        "The 'Correction / note' column accepts free text. Markdown, bullet points, statutory citations — anything that helps us apply the change. Don't worry about being pretty; we read every line.",
        "",
        "Overall sign-off",
        "At the bottom of every skill sheet there is an 'Overall sign-off' block:",
        "  • Verifier name (your full name)",
        "  • Credential (e.g. CPA Texas, CA(SA), ACA, Steuerberater, EA)",
        "  • Date signed",
        "  • Overall assessment (a sentence or two)",
        "  • Confidence (1–5 scale, 5 = stake-my-reputation)",
        "",
        "Fill these in for every skill you signed off. If you only reviewed part of a skill, mark the unreviewed sections 'skip' and still sign off the part you did.",
        "",
        "After you send the workbook back",
        "We parse the file in roughly an hour, apply your corrections to the markdown body of each skill in Supabase, bump those skills to Tier 1, add your UUID to the jurisdiction_verifications.verified_skill_ids array, and register you on openaccountants.com/network with your photo (if you sent one), Calendly link, and bio.",
        "",
        "Any AI client that has the OpenAccountants connector installed will, from that point on, see your name in the skill provenance footer when it uses a skill you signed off — and will route any /openaccountants:review hand-off for that scope to your Calendly.",
        "",
        "Questions: info@openaccountants.com.",
    ]
    for i, txt in enumerate(rows, start=2):
        cell = ws.cell(row=i, column=1, value=txt)
        cell.alignment = WRAP
        if txt and txt[0].isupper() and ":" not in txt and txt.endswith("?") is False and not txt.startswith(("  ", "•")):
            # Section heading lines
            cell.font = Font(bold=True, size=12)
            cell.fill = SUBHEAD_FILL
        ws.row_dimensions[i].height = max(20, 16 * (len(txt) // 100 + 1))


def build_summary_sheet(wb, skills):
    ws = wb.create_sheet("Summary", 2)
    headers = ["#", "Slug", "Title", "Jurisdiction", "Category", "Current tier", "# sections", "Your status", "One-line comment"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    widths = [5, 42, 40, 14, 18, 14, 12, 22, 60]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    summary_dv = DataValidation(
        type="list",
        formula1='"signed,partial,needs-rewrite,not-yet-reviewed"',
        allow_blank=True,
    )
    summary_dv.prompt = "signed = full sign-off • partial = some sections only • needs-rewrite = ask for a re-research pass • not-yet-reviewed = leave blank or this"
    summary_dv.promptTitle = "Skill-level status"
    ws.add_data_validation(summary_dv)

    for i, s in enumerate(skills, start=2):
        n_sections = len(parse_sections(s["markdown"]))
        ws.cell(row=i, column=1, value=i - 1)
        ws.cell(row=i, column=2, value=s["slug"])
        ws.cell(row=i, column=3, value=s.get("name") or s["slug"])
        ws.cell(row=i, column=4, value=s.get("jurisdiction"))
        ws.cell(row=i, column=5, value=s.get("category"))
        ws.cell(row=i, column=6, value=f"Tier {s.get('tier')}")
        ws.cell(row=i, column=7, value=n_sections)
        ws.cell(row=i, column=8).value = ""
        summary_dv.add(f"H{i}")
        ws.cell(row=i, column=9, value="")
        if i % 2 == 0:
            for col in range(1, len(headers) + 1):
                ws.cell(row=i, column=col).fill = ZEBRA_FILL


def build_skill_sheet(wb, s):
    ws = wb.create_sheet(sheet_name(s["slug"]))

    # Column widths chosen for readability
    widths = [38, 90, 22, 60, 28]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Title block
    ws["A1"] = s["slug"]
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:E1")

    ws["A2"] = s.get("name") or ""
    ws["A2"].font = Font(italic=True, color="666666", size=11)
    ws.merge_cells("A2:E2")

    meta_rows = [
        ("Jurisdiction", s.get("jurisdiction")),
        ("Category", s.get("category")),
        ("Current tier", s.get("tier")),
        ("Tax year", s.get("tax_year")),
        ("Source", f"openaccountants.com/skills/{s['slug']}"),
    ]
    for i, (k, v) in enumerate(meta_rows, start=3):
        ws.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws.cell(row=i, column=2, value=str(v) if v is not None else "—")

    # Description
    head_row = 3 + len(meta_rows) + 1
    ws.cell(row=head_row, column=1, value="Description").font = Font(bold=True)
    ws.cell(row=head_row, column=2, value=s.get("description") or "").alignment = WRAP
    ws.merge_cells(start_row=head_row, start_column=2, end_row=head_row, end_column=5)
    ws.row_dimensions[head_row].height = 80

    # Section review table header
    table_head_row = head_row + 2
    headers = ["Section heading", "Section content (truncated to 30k chars)", "Status", "Correction / note", "Confidence (1–5)"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=table_head_row, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(vertical="center")
    ws.row_dimensions[table_head_row].height = 22
    ws.freeze_panes = f"A{table_head_row + 1}"

    # Per-section rows
    status_dv = DataValidation(
        type="list",
        formula1=STATUS_OPTIONS,
        allow_blank=True,
    )
    status_dv.prompt = STATUS_DV.prompt
    status_dv.promptTitle = STATUS_DV.promptTitle
    ws.add_data_validation(status_dv)

    confidence_dv = DataValidation(
        type="list",
        formula1='"1,2,3,4,5"',
        allow_blank=True,
    )
    ws.add_data_validation(confidence_dv)

    sections = parse_sections(s["markdown"])
    row = table_head_row + 1
    for level, heading, content in sections:
        prefix = "  " * (level - 1)
        ws.cell(row=row, column=1, value=f"{prefix}{heading}").alignment = WRAP
        # Cell text limit ~32k. Truncate longer sections with a note.
        snippet = content if len(content) <= 30000 else (content[:30000] + "\n\n[…truncated for cell limit; see github.com/openaccountants/openaccountants for full content]")
        ws.cell(row=row, column=2, value=snippet).alignment = WRAP
        ws.cell(row=row, column=3).alignment = TOP
        status_dv.add(f"C{row}")
        ws.cell(row=row, column=4).alignment = WRAP
        ws.cell(row=row, column=5).alignment = TOP
        confidence_dv.add(f"E{row}")
        # Estimate row height by content length
        rough_lines = max(snippet.count("\n") + 1, len(snippet) // 100 + 1)
        ws.row_dimensions[row].height = min(400, max(30, rough_lines * 14))
        if row % 2 == 0:
            for col in range(1, 6):
                ws.cell(row=row, column=col).fill = ZEBRA_FILL
        row += 1

    # Sign-off block at the bottom
    row += 2
    so_header = ws.cell(row=row, column=1, value="OVERALL SIGN-OFF")
    so_header.font = Font(bold=True, size=12, color="9A3412")
    so_header.fill = SIGNOFF_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)

    signoff_fields = [
        ("Verifier name", args.verifier),
        ("Credential (CPA Texas, CA(SA), ACA, Steuerberater, EA, etc.)", args.credential),
        ("Date signed", ""),
        ("Overall assessment (one or two sentences)", ""),
        ("Confidence in this skill as a whole (1–5)", ""),
        ("Any sections you didn't have time to review (slugs / headings)", ""),
        ("Suggested re-research topics (if any)", ""),
    ]
    for k, v in signoff_fields:
        row += 1
        ws.cell(row=row, column=1, value=k).font = Font(bold=True)
        ws.cell(row=row, column=1).fill = SIGNOFF_FILL
        ws.cell(row=row, column=2, value=v).alignment = WRAP
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
        ws.row_dimensions[row].height = 24


if __name__ == "__main__":
    main()
